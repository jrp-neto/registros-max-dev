import logs
import time
import win32com.client
import re
import logs
import pythoncom
import webbrowser
import xlwings as xw
import pywintypes
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException

def login(user: str, password: str, chrome: webdriver.Chrome):
    # Preencher usuário
    element_user = WebDriverWait(chrome, 30).until(EC.presence_of_element_located((By.XPATH,"/html/body/form/div[3]/div[1]/div[1]/div[2]/div[1]/input")))
    element_user.send_keys(f"SEREDUC\\{str(user)}")
    # Preencher senha
    element_password = WebDriverWait(chrome, 30).until(EC.presence_of_element_located((By.XPATH,"/html/body/form/div[3]/div[1]/div[1]/div[2]/div[2]/input")))
    element_password.send_keys(password)
    # Fazer login e acessar a página de buscar discente
    element_login = WebDriverWait(chrome, 30).until(EC.element_to_be_clickable((By.XPATH,"/html/body/form/div[3]/div[1]/div[1]/div[2]/div[3]/div[2]/input")))
    chrome.execute_script("arguments[0].click();", element_login)
    try:
        if WebDriverWait(chrome, 5).until(EC.presence_of_element_located((By.ID,"SilkUIFramework_wt7_block_wt6_RichWidgets_wt11_block_wtSanitizedHtml2"))):
            logs.logging.info("Usuário ou senha estão incorretos!")
            return True
    except:
        logs.logging.info(f"Login realizado com sucesso no usuário: {user}")

def search(enrolment: str, chrome: webdriver.Chrome) -> None:
    # Preencher a matrícula na busca
    element_enrolment = WebDriverWait(chrome, 30).until(EC.presence_of_element_located((By.XPATH,"/html/body/form/div[3]/div[3]/div[1]/div[2]/div[1]/div/div/div[1]/div/div/div[1]/div/div/input")))
    element_enrolment.send_keys(str(enrolment).zfill(8))
    # Alterar o tipo da busca para matrícula
    element_search_type = WebDriverWait(chrome, 30).until(EC.presence_of_element_located((By.XPATH,"/html/body/form/div[3]/div[3]/div[1]/div[2]/div/div/div/div[1]/div/div/div[2]/div/select")))
    element_search_type.send_keys("Mat")
    # Buscar discente
    element_search = WebDriverWait(chrome, 30).until(EC.element_to_be_clickable((By.XPATH,"/html/body/form/div[3]/div[3]/div[1]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div/input")))
    element_search.click()
    # Selecionar discente no resultado da busca e redirecionar para a tela de registros
    elemento_student = WebDriverWait(chrome, 30).until(EC.element_to_be_clickable((By.XPATH,"/html/body/form/div[3]/div[3]/div[1]/div[2]/table/tbody/tr/td[4]/div/a")))
    elemento_student.click()

def register(registry: str, chrome: webdriver.Chrome) -> None:
    # Abrir o "frame" de registro em uma nova aba
    element_create_occurrence = WebDriverWait(chrome, 30).until(EC.presence_of_element_located((By.ID,"DublinTheme_wt66_block_wtMainContent_SilkUIFramework_wt266_block_wtColumn1_SilkUIFramework_wtFiltroContainer_block_wtActions_wtCreateOcorrencia")))
    chrome.execute_script(f"window.open('{element_create_occurrence.get_attribute('href')}', '_blank');")
    chrome.switch_to.window(chrome.window_handles[1])
    # Selecionar a opção do primeiro menu dropdown
    dropdown_1 = WebDriverWait(chrome, 30).until(EC.element_to_be_clickable((By.XPATH,"/html/body/form/div[3]/div/div/div[1]/div/div[1]/div/a[1]/span[2]/b")))
    dropdown_1.click()
    input_1 = WebDriverWait(chrome, 30).until(EC.presence_of_element_located((By.XPATH,"/html/body/div[2]/div/input")))
    input_1.send_keys("Tut" + Keys.ENTER)
    time.sleep(0.5)
    # Selecionar a opção do segundo menu dropdown
    for _ in range(2):
        try:
            dropdown_2 = WebDriverWait(chrome, 30).until(EC.element_to_be_clickable((By.XPATH,"/html/body/form/div[3]/div/div/div[1]/div/div[2]/div/a[1]/span[2]/b")))
            dropdown_2.click()
            input_2 = WebDriverWait(chrome, 30).until(EC.presence_of_element_located((By.XPATH,"/html/body/div[3]/div/input")))
            input_2.send_keys("Ori" + Keys.ENTER)
            time.sleep(0.25)
        except StaleElementReferenceException:
            logs.logging.info('Não foi possível selecionar a opção "Orientações Gerais", tentando novamente...')
            chrome.close()
            chrome.switch_to.window(chrome.window_handles[0])
            register(registry, chrome)
            return
    # Selecionar a opção do terceiro menu dropdown
    dropdown_3 = WebDriverWait(chrome, 30).until(EC.element_to_be_clickable((By.XPATH,"/html/body/form/div[3]/div/div/div[1]/div/div[3]/select")))
    dropdown_3.send_keys(Keys.END)
    time.sleep(0.5)
    # Preencher registro
    element_obs = WebDriverWait(chrome, 30).until(EC.presence_of_element_located((By.XPATH,"/html/body/form/div[3]/div/div/div[1]/div/div[5]/textarea")))
    element_obs.send_keys(registry)
    # Salvar o registro
    element_save = WebDriverWait(chrome, 30).until(EC.element_to_be_clickable((By.XPATH,"/html/body/form/div[3]/div/div/div[2]/a")))
    element_save.click()
    # Fecha a aba de registro, voltando para a aba de registros do discente atual
    time.sleep(2)
    chrome.close()
    chrome.switch_to.window(chrome.window_handles[0])

def verify(registry: str, enrolment: str, chrome: webdriver.Chrome) -> None:
    # Atualizar a página e aguardar carregamento completo para verificar se o registro foi salvo
    time.sleep(1)
    chrome.refresh()
    while chrome.execute_script("return document.readyState") != "complete":
        time.sleep(0.1)
    time.sleep(1)
    # Verificar se o registro foi realizado e caso não tenha sido, tentar novamente
    element_register_1 = WebDriverWait(chrome, 30).until(EC.presence_of_element_located((By.XPATH,"/html/body/form/div[3]/div[3]/div[1]/div[2]/div[2]/div/div/div[2]/div[1]/div/span/table/tbody/tr[1]/td[5]/div")))
    element_register_2 = WebDriverWait(chrome, 30).until(EC.presence_of_element_located((By.XPATH,"/html/body/form/div[3]/div[3]/div[1]/div[2]/div[2]/div/div/div[2]/div[1]/div/span/table/tbody/tr[2]/td[5]/div")))
    if element_register_1.text == registry or element_register_2.text == registry:
        return
    else:
        logs.logging.info(f"Registro na matrícula {str(enrolment).zfill(8)} não foi salvo. Uma nova tentativa será realizada!")
        register(registry, chrome)
        verify(registry, enrolment, chrome)

def extract(folder_name: str, spreadsheet: str):
    try:
        # Conectar ao Outlook
        pythoncom.CoInitialize()
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        # Sincronizar Outlook
        namespace.SyncObjects.Item(1).Start
        time.sleep(30)
        # Acessar a conta de e-mail principal
        email_account = namespace.Folders(1)
        # Acessar a pasta desejada
        try:
            folder = email_account.Folders[folder_name]
        except:
            inbox = namespace.GetDefaultFolder(6)
            folder = inbox.Folders[folder_name]
        # Filtrar e-mails não lidos
        unread_messages = folder.Items.Restrict("[Unread] = true")
        # Ordenar os e-mails do mais recente para o mais antigo
        unread_messages.Sort("[ReceivedTime]", True)
        # Criar uma lista para armazenar os e-mails e assuntos
        emails_info = []
        for message in unread_messages:
            if message.Class == 43:  # 43 = MailItem (ignora reuniões, tarefas, etc.)
                # Pegar o endereço de e-mail do remetente
                email_sender = message.Sender.GetExchangeUser().PrimarySmtpAddress if message.SenderEmailType == "EX" else message.SenderEmailAddress
                # Pegar o assunto do e-mail
                subject = message.Subject
                # Remover prefixos comuns de respostas automáticas ou e-mails de resposta
                subject = re.sub(r'^(RES:|Res:|RE:|Re:|RESPOSTA AUTOMÁTICA:|Resposta automática:|ENC:|Enc:|ASSUNTO:|Assunto:|FÉRIAS RE:|Férias Re:)\s*', '', subject).strip()
                # Adicionar à lista
                emails_info.append([email_sender, subject])
        # Filtrar erros
        block_senders = {
            "postmaster@outlook.com",
            "postmaster@sesc.com.br",
            "MicrosoftExchange329e71ec88ae4615bbc36ab6ce41109e@SEREDUCACIONALBR.onmicrosoft.com",
            "MAILER-DAEMON@mx3.bol.com.br"
        }
        emails_info = [info for info in emails_info if info[0].lower() not in block_senders]
        # Carrgar o arquivo Excel na aba correta
        wb = load_workbook(spreadsheet)
        ws = wb["Registros CRA - Ações"]
        # Limpar apenas as colunas A e B
        for row in ws.iter_rows(min_col=1, max_col=2):
            for cell in row:
                cell.value = None
        # Adicionar cabeçalho na primeira linha
        ws["A1"] = "E-mail"
        ws["B1"] = "Assunto"
        # Adicionar os novos dados de E-mails e Assuntos
        for index, (email_remetente, assunto) in enumerate(emails_info, start=2):
            ws[f"A{index}"] = email_remetente
            ws[f"B{index}"] = assunto
        # Salvar arquivo Excel
        wb.save(spreadsheet)
        wb.close()
        # Abre o arquivo Excel em segundo plano para atualizar as fórmulas
        app = xw.App(visible=False)
        wb = app.books.open(spreadsheet)
        wb.app.calculate()
        wb.save()
        wb.close()
        app.quit()
        logs.logging.info("Extração finalizada!")
        return "Extração finalizada!", "green", 60000
    except PermissionError:
        logs.logging.info("Feche a planilha antes de inciar a extração!")
        return "Feche a planilha antes de inciar a extração!", "red", 10000
    except FileNotFoundError:
        logs.logging.info("Arquivo Excel não encontrado!")
        return "Arquivo Excel não encontrado!", "red", 10000
    except pywintypes.com_error:
        logs.logging.info("Pasta não encontrada!")
        return "Pasta não encontrada!", "red", 10000
    finally:
        pythoncom.CoUninitialize()

def abrir_link(event):
    webbrowser.open("https://github.com/jrp-neto/registros-max")